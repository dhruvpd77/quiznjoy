from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.core.files.base import ContentFile
from django.urls import reverse
from django.http import HttpResponse
from .models import Semester, Subject, Question, Unit, ProgrammingQuestion, ProgrammingQuestionAccess, QuestionReport
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta
import pandas as pd
import openpyxl
from io import BytesIO
import os
import csv
from PIL import Image

@staff_member_required
def admin_dashboard(request):
    semesters = Semester.objects.all()
    subjects = Subject.objects.all()
    questions_count = Question.objects.count()
    
    context = {
        'semesters': semesters,
        'subjects': subjects,
        'questions_count': questions_count,
    }
    return render(request, 'semesters/admin_dashboard.html', context)

@staff_member_required
def create_semester(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        
        if name:
            Semester.objects.create(name=name, description=description)
            messages.success(request, f'Semester "{name}" created successfully!')
            return redirect('semesters:admin_dashboard')
        else:
            messages.error(request, 'Semester name is required.')
    
    return render(request, 'semesters/create_semester.html')

@staff_member_required
def edit_semester(request, semester_id):
    semester = get_object_or_404(Semester, id=semester_id)
    
    if request.method == 'POST':
        semester.name = request.POST.get('name', semester.name)
        semester.description = request.POST.get('description', semester.description)
        semester.save()
        messages.success(request, f'Semester "{semester.name}" updated successfully!')
        return redirect('semesters:manage_semesters')
    
    return render(request, 'semesters/edit_semester.html', {'semester': semester})

@staff_member_required
def delete_semester(request, semester_id):
    semester = get_object_or_404(Semester, id=semester_id)
    
    if request.method == 'POST':
        semester_name = semester.name
        semester.delete()
        messages.success(request, f'Semester "{semester_name}" deleted successfully!')
        return redirect('semesters:manage_semesters')
    
    # Get counts
    subjects_count = semester.subjects.count()
    questions_count = sum(subject.questions.count() for subject in semester.subjects.all())
    
    return render(request, 'semesters/delete_semester.html', {
        'semester': semester,
        'subjects_count': subjects_count,
        'questions_count': questions_count,
    })

@staff_member_required
def edit_subject(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    semesters = Semester.objects.all()
    
    if request.method == 'POST':
        subject.name = request.POST.get('name', subject.name)
        subject.code = request.POST.get('code', subject.code)
        semester_id = request.POST.get('semester')
        if semester_id:
            subject.semester = get_object_or_404(Semester, id=semester_id)
        subject.save()
        messages.success(request, f'Subject "{subject.name}" updated successfully!')
        return redirect('semesters:manage_subjects', semester_id=subject.semester.id)
    
    return render(request, 'semesters/edit_subject.html', {
        'subject': subject,
        'semesters': semesters
    })

@staff_member_required
def delete_subject(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    semester_id = subject.semester.id
    
    if request.method == 'POST':
        subject_name = subject.name
        subject.delete()
        messages.success(request, f'Subject "{subject_name}" deleted successfully!')
        return redirect('semesters:manage_subjects', semester_id=semester_id)
    
    questions_count = subject.questions.count()
    
    return render(request, 'semesters/delete_subject.html', {
        'subject': subject,
        'questions_count': questions_count,
    })

@staff_member_required
def create_subject(request):
    semesters = Semester.objects.all()
    
    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code', '')
        semester_id = request.POST.get('semester')
        
        if name and semester_id:
            semester = get_object_or_404(Semester, id=semester_id)
            Subject.objects.create(name=name, code=code, semester=semester)
            messages.success(request, f'Subject "{name}" created successfully!')
            return redirect('semesters:admin_dashboard')
        else:
            messages.error(request, 'Subject name and semester are required.')
    
    return render(request, 'semesters/create_subject.html', {'semesters': semesters})

@staff_member_required
def download_upload_template(request):
    """Download sample Excel template for question upload. Compatible with all subjects: images (e.g. Digital Electronics gates), code/indent (Full Stack), math formulas."""
    # Use column names that match upload parser: option A, option B, etc. (with space)
    df = pd.DataFrame([
        {
            'unit_number': 1,
            'question_text': 'Sample question text? (Replace with your question)',
            'MCQ Answer': 'A',
            'option A': 'Option A text',
            'option B': 'Option B text',
            'option C': 'Option C text',
            'option D': 'Option D text',
            'Added By': '',
            'Verified By': '',
        },
        {
            'unit_number': 1,
            'question_text': 'For code/Full Stack: use Alt+Enter in cell for new lines; keep spaces for indentation.\nExample:\ndef hello():\n    print("Hi")',
            'MCQ Answer': 'B',
            'option A': 'Syntax error',
            'option B': 'Prints Hi',
            'option C': 'None',
            'option D': 'Prints hello',
            'Added By': '',
            'Verified By': '',
        },
        {
            'unit_number': 2,
            'question_text': 'Math formulas: use \\( and \\) for inline math, \\[ and \\] for display. Example: \\( x^2 + y^2 = z^2 \\)',
            'MCQ Answer': 'A',
            'option A': '\\( x^2 + y^2 = z^2 \\)',
            'option B': 'x squared plus y squared',
            'option C': 'Pythagoras',
            'option D': 'None of these',
            'Added By': '',
            'Verified By': '',
        },
    ])
    buf = BytesIO()
    df.to_excel(buf, index=False, engine='openpyxl')
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="questions_upload_template.xlsx"'
    return response


@staff_member_required
def upload_questions(request):
    subjects = Subject.objects.all()
    
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        subject_id = request.POST.get('subject')
        
        if not excel_file:
            messages.error(request, 'Please select an Excel file.')
            return render(request, 'semesters/upload_questions.html', {'subjects': subjects})
        
        if not subject_id:
            messages.error(request, 'Please select a subject.')
            return render(request, 'semesters/upload_questions.html', {'subjects': subjects})
        
        subject = get_object_or_404(Subject, id=subject_id)
        
        try:
            # Save uploaded file temporarily
            file_path = f'/tmp/{excel_file.name}' if os.name != 'nt' else f'C:\\temp\\{excel_file.name}'
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            with open(file_path, 'wb+') as destination:
                for chunk in excel_file.chunks():
                    destination.write(chunk)
            
            # Try to extract images, but continue even if it fails
            images_dict = {}
            try:
                # Load workbook with openpyxl to extract images
                # Use data_only=True to avoid formula issues, read_only=False for images
                wb = openpyxl.load_workbook(file_path, data_only=True, keep_vba=False)
                ws = wb.active
                
                # Extract all images from the worksheet and map them to cells
                if hasattr(ws, '_images'):
                    for image in ws._images:
                        # Get the cell anchor (where image is placed)
                        try:
                            if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                                row_idx = image.anchor._from.row + 1  # openpyxl is 0-indexed
                                col_idx = image.anchor._from.col + 1
                                
                                # Save image to BytesIO
                                img_data = BytesIO(image._data())
                                images_dict[row_idx] = img_data
                        except Exception as img_error:
                            # Skip this image if there's an error
                            continue
            except Exception as openpyxl_error:
                # If openpyxl fails, continue without images
                pass  # Will show message later
            
            # Try multiple engines to read the Excel file
            df = None
            engines_to_try = ['openpyxl', 'xlrd', None]  # None lets pandas choose
            last_error = None
            
            for engine in engines_to_try:
                try:
                    if engine:
                        df = pd.read_excel(file_path, dtype=str, keep_default_na=False, engine=engine)
                    else:
                        df = pd.read_excel(file_path, dtype=str, keep_default_na=False)
                    break  # Success! Exit loop
                except Exception as e:
                    last_error = e
                    continue  # Try next engine
            
            # If pandas failed, try reading cell-by-cell with openpyxl directly (LAST RESORT)
            if df is None:
                # Try multiple openpyxl strategies
                openpyxl_strategies = [
                    # Strategy 1: read_only mode (most forgiving)
                    {'read_only': True, 'data_only': True, 'keep_vba': False},
                    # Strategy 2: read_only without data_only
                    {'read_only': True, 'data_only': False, 'keep_vba': False},
                    # Strategy 3: Normal mode with repair
                    {'read_only': False, 'data_only': True, 'keep_vba': False},
                    # Strategy 4: Normal mode without data_only
                    {'read_only': False, 'data_only': False, 'keep_vba': False},
                ]
                
                for strategy_idx, strategy in enumerate(openpyxl_strategies):
                    try:
                        if strategy_idx == 0:
                            messages.warning(request, 'File has severe XML corruption. Attempting direct cell-by-cell reading...')
                        
                        # Try to open with current strategy
                        wb_readonly = openpyxl.load_workbook(file_path, **strategy)
                        ws_readonly = wb_readonly.active
                        
                        # Read data manually row by row
                        data_rows = []
                        headers = None
                        max_rows = 10000  # Safety limit
                        
                        for idx, row in enumerate(ws_readonly.iter_rows(values_only=True)):
                            if idx >= max_rows:
                                break
                            
                            if idx == 0:
                                # First row is headers
                                headers = [str(cell) if cell is not None else '' for cell in row]
                            else:
                                # Data rows
                                row_data = [str(cell) if cell is not None else '' for cell in row]
                                if any(row_data):  # Skip completely empty rows
                                    data_rows.append(row_data)
                        
                        # Create DataFrame from manually read data
                        if headers and data_rows:
                            df = pd.DataFrame(data_rows, columns=headers)
                            messages.success(request, f'Successfully read {len(df)} rows using cell-by-cell method (strategy {strategy_idx + 1})!')
                            wb_readonly.close()
                            break  # Success! Exit loop
                        
                        wb_readonly.close()
                    
                    except Exception as cell_error:
                        # Try next strategy
                        continue
                
                # If all openpyxl strategies failed, try CSV conversion as absolute last resort
                if df is None:
                    try:
                        messages.warning(request, 'All Excel reading methods failed. Attempting CSV conversion workaround...')
                        
                        # Try to read as CSV (if file can be opened as text)
                        csv_data = []
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            # Try to detect if it's actually a CSV or TSV
                            sample = f.read(1024)
                            f.seek(0)
                            
                            # Try comma-separated first
                            try:
                                reader = csv.reader(f)
                                for row in reader:
                                    if row:  # Skip empty rows
                                        csv_data.append(row)
                                if len(csv_data) > 1:  # At least header + 1 row
                                    headers = [str(h) for h in csv_data[0]]
                                    data_rows = [row for row in csv_data[1:] if any(row)]
                                    if headers and data_rows:
                                        df = pd.DataFrame(data_rows, columns=headers)
                                        messages.success(request, f'Successfully read {len(df)} rows using CSV fallback method!')
                            except:
                                pass
                    except Exception as csv_error:
                        pass
            
            # If still no data, provide helpful error message with actionable steps
            if df is None:
                # Create a more user-friendly error message
                error_msg = (
                    "❌ Cannot read Excel file due to severe XML corruption.\n\n"
                    "🔧 QUICK FIX (Recommended - Takes 2 minutes):\n"
                    "1. Open the corrupted file in Microsoft Excel\n"
                    "2. Press Ctrl+A to select all data\n"
                    "3. Press Ctrl+C to copy\n"
                    "4. Close the corrupted file\n"
                    "5. Create a NEW blank Excel workbook\n"
                    "6. Press Ctrl+V to paste\n"
                    "7. Save as new .xlsx file (e.g., 'Fixed_File.xlsx')\n"
                    "8. Upload the new file\n\n"
                    "💡 Alternative: Save as CSV first, then convert back to Excel\n"
                    "1. Open in Excel → File → Save As → CSV (Comma delimited)\n"
                    "2. Close and reopen the CSV file\n"
                    "3. Save As → Excel Workbook (.xlsx)\n"
                    "4. Upload the new file\n\n"
                    "🛠️ Or run fix script: python fix_corrupted_excel.py \"your_file.xlsx\"\n\n"
                    "⚠️ The file structure is too corrupted for automatic repair."
                )
                raise Exception(error_msg)
            
            # Expected columns:
            # unit_number | question_text | MCQ Answer | option A | option B | option C | option D | Added By | Verified By
            
            # Normalize column names: strip whitespace and accept common Excel variants (e.g. "unit_numb")
            df.columns = [str(c).strip() for c in df.columns]
            if 'unit_numb' in df.columns and 'unit_number' not in df.columns:
                df = df.rename(columns={'unit_numb': 'unit_number'})
            
            def has_question_column(dataf):
                for c in dataf.columns:
                    if 'question' in str(c).lower():
                        return True
                return False
            
            # If no question-like column, header might be in row 2 (many Excel files have title in row 1)
            if not has_question_column(df) and len(df) >= 2:
                for engine in ['openpyxl', 'xlrd', None]:
                    try:
                        if engine:
                            df_retry = pd.read_excel(file_path, dtype=str, keep_default_na=False, engine=engine, header=1)
                        else:
                            df_retry = pd.read_excel(file_path, dtype=str, keep_default_na=False, header=1)
                        df_retry.columns = [str(c).strip() for c in df_retry.columns]
                        if 'unit_numb' in df_retry.columns and 'unit_number' not in df_retry.columns:
                            df_retry = df_retry.rename(columns={'unit_numb': 'unit_number'})
                        if has_question_column(df_retry):
                            df = df_retry.reset_index(drop=True)
                            break
                    except Exception:
                        continue
            
            # Debug: Print column names to help diagnose issues
            print(f"DEBUG: DataFrame columns: {list(df.columns)}")
            print(f"DEBUG: Total rows in DataFrame: {len(df)}")
            if len(df) > 0:
                print(f"DEBUG: First row sample: {df.iloc[0].to_dict()}")
            
            questions_created = 0
            programming_questions_created = 0
            skipped_count = 0
            skipped_rows = []
            images_extracted = 0
            
            for index, row in df.iterrows():
                try:
                    # Excel row number (pandas index + 2 for header + 0-indexing)
                    excel_row = index + 2
                    
                    # Extract unit number (accept "unit_number" or "unit_numb")
                    unit = row.get('unit_number') or row.get('unit_numb', '1')
                    if not unit or unit == '' or unit == 'nan':
                        unit = 1
                    else:
                        try:
                            unit = int(float(unit))
                        except:
                            unit = 1
                    
                    # Get values: strip only leading/trailing whitespace so newlines and indentation
                    # (e.g. for Full Stack code, Digital Electronics) are preserved in question_text and options.
                    def get_row_value(row, col_name, default=''):
                        """Get value from pandas Series row; preserves internal newlines/indent for code and text."""
                        val = None
                        if col_name in row:
                            val = row[col_name]
                        else:
                            for key in row.index:
                                if str(key).strip().lower() == str(col_name).strip().lower():
                                    val = row[key]
                                    break
                        if val is None:
                            return default
                        try:
                            if pd.isna(val):
                                return default
                        except Exception:
                            pass
                        if val is None or str(val).strip().lower() in ['nan', 'none', 'nat', '']:
                            return default
                        # .strip() only removes leading/trailing; internal newlines/spaces kept for code/indent
                        return str(val).strip() if val is not None else default
                    
                    # Get column values with flexible matching
                    question_text = get_row_value(row, 'question_text') or get_row_value(row, 'question text') or get_row_value(row, 'Question')
                    mcq_answer_raw = str(get_row_value(row, 'MCQ Answer') or get_row_value(row, 'mcq answer') or get_row_value(row, 'Answer') or '').strip()
                    mcq_answer = mcq_answer_raw.upper()  # Uppercase for letter matching
                    option_a = get_row_value(row, 'option A') or get_row_value(row, 'option_a') or get_row_value(row, 'A')
                    option_b = get_row_value(row, 'option B') or get_row_value(row, 'option_b') or get_row_value(row, 'B')
                    option_c = get_row_value(row, 'option C') or get_row_value(row, 'option_c') or get_row_value(row, 'C')
                    option_d = get_row_value(row, 'option D') or get_row_value(row, 'option_d') or get_row_value(row, 'D')
                    solution = get_row_value(row, 'solution') or get_row_value(row, 'Solution') or ''
                    added_by = get_row_value(row, 'Added By') or get_row_value(row, 'added_by') or ''
                    verified_by = get_row_value(row, 'Verified By') or get_row_value(row, 'verified_by') or ''
                    
                    # Clean 'nan' strings but preserve actual content
                    def clean_value(val):
                        if not val or val == '':
                            return ''
                        val_str = str(val).strip()
                        if val_str.lower() in ['nan', 'none', 'nat', '']:
                            return ''
                        return val_str
                    
                    # Clean all values
                    question_text = clean_value(question_text)
                    option_a = clean_value(option_a)
                    option_b = clean_value(option_b)
                    option_c = clean_value(option_c)
                    option_d = clean_value(option_d)
                    solution = clean_value(solution)
                    
                    # Skip if question text is empty
                    if not question_text or question_text.strip() == '':
                        skipped_count += 1
                        skipped_rows.append((excel_row, 'empty question text'))
                        continue
                    
                    # Check if all 4 options are present (MCQ) or not (Programming Question)
                    has_all_options = (
                        option_a and option_a.strip() != '' and
                        option_b and option_b.strip() != '' and
                        option_c and option_c.strip() != '' and
                        option_d and option_d.strip() != ''
                    )
                    
                    if not has_all_options:
                        # This is a programming question - save it as ProgrammingQuestion
                        programming_question = ProgrammingQuestion.objects.create(
                            subject=subject,
                            unit=unit,
                            question_text=question_text,
                            solution=clean_value(solution),
                            added_by=clean_value(added_by),
                            verified_by=clean_value(verified_by)
                        )
                        
                        # Check if this row has an image
                        if excel_row in images_dict:
                            try:
                                img_data = images_dict[excel_row]
                                img_data.seek(0)
                                
                                # Generate unique filename
                                filename = f'prog_question_{programming_question.id}_{excel_row}.png'
                                
                                # Save image to question
                                programming_question.question_image.save(filename, ContentFile(img_data.read()), save=True)
                                images_extracted += 1
                            except Exception as img_error:
                                pass  # Continue even if image extraction fails
                        
                        programming_questions_created += 1
                        continue  # Skip MCQ processing for programming questions
                    
                    # If we reach here, it's a complete MCQ with all 4 options
                    
                    # Convert MCQ Answer to letter format (A, B, C, D)
                    # Support both formats:
                    # 1. Existing format: "A", "B", "C", "D", "Option A", "A)", etc.
                    # 2. New format: Actual text from options (e.g., "Uniform Resource Locator")
                    correct_answer = None
                    
                    # Helper function to normalize text for comparison (case-insensitive, normalize whitespace)
                    def normalize_text(text):
                        if not text:
                            return ''
                        # Strip and normalize multiple spaces to single space
                        normalized = ' '.join(str(text).strip().split())
                        return normalized
                    
                    # First, try existing format (letter-based matching)
                    if mcq_answer in ['A', 'B', 'C', 'D']:
                        correct_answer = mcq_answer
                    elif mcq_answer.startswith('A'):
                        correct_answer = 'A'
                    elif mcq_answer.startswith('B'):
                        correct_answer = 'B'
                    elif mcq_answer.startswith('C'):
                        correct_answer = 'C'
                    elif mcq_answer.startswith('D'):
                        correct_answer = 'D'
                    
                    # If not found, try new format (text-based matching)
                    # Compare the MCQ answer text with each option text (case-insensitive, normalized whitespace)
                    if correct_answer is None:
                        mcq_answer_normalized = normalize_text(mcq_answer_raw).lower()
                        option_a_normalized = normalize_text(option_a).lower()
                        option_b_normalized = normalize_text(option_b).lower()
                        option_c_normalized = normalize_text(option_c).lower()
                        option_d_normalized = normalize_text(option_d).lower()
                        
                        # Case-insensitive exact match
                        if mcq_answer_normalized == option_a_normalized:
                            correct_answer = 'A'
                        elif mcq_answer_normalized == option_b_normalized:
                            correct_answer = 'B'
                        elif mcq_answer_normalized == option_c_normalized:
                            correct_answer = 'C'
                        elif mcq_answer_normalized == option_d_normalized:
                            correct_answer = 'D'
                    
                    # Default fallback if no match found
                    if correct_answer is None:
                        correct_answer = 'A'  # Default
                    
                    # Create question
                    question = Question.objects.create(
                        subject=subject,
                        unit=unit,
                        question_text=question_text,
                        option_a=option_a,
                        option_b=option_b,
                        option_c=option_c,
                        option_d=option_d,
                        correct_answer=correct_answer,
                        added_by=clean_value(added_by),
                        verified_by=clean_value(verified_by)
                    )
                    
                    # Check if this row has an image
                    if excel_row in images_dict:
                        try:
                            img_data = images_dict[excel_row]
                            img_data.seek(0)
                            
                            # Generate unique filename
                            filename = f'question_{question.id}_{excel_row}.png'
                            
                            # Save image to question
                            question.question_image.save(filename, ContentFile(img_data.read()), save=True)
                            images_extracted += 1
                        except Exception as img_error:
                            pass  # Continue even if image extraction fails
                    
                    questions_created += 1
                except Exception as e:
                    import traceback
                    print(f"ERROR processing row {excel_row}: {traceback.format_exc()}")
                    skipped_count += 1
                    skipped_rows.append((excel_row, str(e)[:80]))
                    continue
            
            # Clean up temp file
            try:
                os.remove(file_path)
            except:
                pass
            
            # Build success message
            msg_parts = []
            if questions_created > 0:
                msg_parts.append(f'✅ {questions_created} MCQ questions')
            if programming_questions_created > 0:
                msg_parts.append(f'✅ {programming_questions_created} Programming questions')
            
            if questions_created > 0 or programming_questions_created > 0:
                msg = 'Uploaded successfully: ' + ', '.join(msg_parts) + '!'
                if images_extracted > 0:
                    msg += f' ({images_extracted} questions with images)'
                elif len(images_dict) == 0:
                    msg += ' ⚠️ No images were extracted (file may have XML issues or no images present)'
                if skipped_count > 0:
                    msg += f'. Skipped {skipped_count} row(s): '
                    msg += ', '.join(f'row {r} ({reason})' for r, reason in skipped_rows[:10])
                    if len(skipped_rows) > 10:
                        msg += f' ... and {len(skipped_rows)-10} more'
                messages.success(request, msg)
            else:
                messages.warning(request, f'No valid questions found. Skipped {skipped_count} rows. Make sure questions have text, and MCQs have all 4 options (A, B, C, D).')
            return redirect('semesters:admin_dashboard')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            # Clean up temp file on error
            try:
                if 'file_path' in locals():
                    os.remove(file_path)
            except:
                pass
    
    return render(request, 'semesters/upload_questions.html', {'subjects': subjects})


@staff_member_required
def export_questions(request, subject_id):
    """Export MCQ questions for a subject to Excel"""
    subject = get_object_or_404(Subject, id=subject_id)
    questions = Question.objects.filter(subject=subject).order_by('unit', 'id')
    rows = []
    for q in questions:
        rows.append({
            'unit_number': q.unit,
            'question_text': q.question_text,
            'MCQ Answer': q.correct_answer,
            'A': q.option_a,
            'B': q.option_b,
            'C': q.option_c,
            'D': q.option_d,
            'explanation': q.explanation or '',
            'hint': q.hint or '',
            'difficulty': q.difficulty or '',
            'tags': q.tags or '',
            'Added By': q.added_by or '',
            'Verified By': q.verified_by or '',
        })
    df = pd.DataFrame(rows)
    buf = BytesIO()
    df.to_excel(buf, index=False, engine='openpyxl')
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    safe_name = "".join(c for c in subject.name if c.isalnum() or c in (' ', '-')).strip()[:50]
    response['Content-Disposition'] = f'attachment; filename="questions_{safe_name}.xlsx"'
    return response


@staff_member_required
def manage_semesters(request):
    semesters = Semester.objects.all()
    return render(request, 'semesters/manage_semesters.html', {'semesters': semesters})

@staff_member_required
def manage_subjects(request, semester_id):
    semester = get_object_or_404(Semester, id=semester_id)
    subjects = Subject.objects.filter(semester=semester)
    return render(request, 'semesters/manage_subjects.html', {
        'semester': semester,
        'subjects': subjects
    })

@staff_member_required
def manage_questions(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    unit_filter = request.GET.get('unit', '')
    
    if unit_filter:
        questions = Question.objects.filter(subject=subject, unit=unit_filter).order_by('id')
    else:
        questions = Question.objects.filter(subject=subject).order_by('unit', 'id')
    
    # Get all available units for this subject
    available_units = Question.objects.filter(subject=subject).values_list('unit', flat=True).distinct().order_by('unit')
    
    context = {
        'subject': subject,
        'questions': questions,
        'available_units': available_units,
        'selected_unit': unit_filter,
    }
    return render(request, 'semesters/manage_questions.html', context)

@staff_member_required
def edit_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    
    if request.method == 'POST':
        question.unit = request.POST.get('unit', question.unit)
        question.question_text = request.POST.get('question_text', question.question_text)
        question.option_a = request.POST.get('option_a', question.option_a)
        question.option_b = request.POST.get('option_b', question.option_b)
        question.option_c = request.POST.get('option_c', question.option_c)
        question.option_d = request.POST.get('option_d', question.option_d)
        question.correct_answer = request.POST.get('correct_answer', question.correct_answer)
        question.explanation = request.POST.get('explanation', question.explanation)
        question.hint = request.POST.get('hint', question.hint)
        question.difficulty = request.POST.get('difficulty', question.difficulty)
        question.tags = request.POST.get('tags', question.tags)
        question.status = request.POST.get('status', question.status)
        question.added_by = request.POST.get('added_by', question.added_by)
        question.verified_by = request.POST.get('verified_by', question.verified_by)
        
        # Handle image upload
        if 'question_image' in request.FILES:
            question.question_image = request.FILES['question_image']
        
        # Handle image removal
        if request.POST.get('remove_image') == 'on' and question.question_image:
            question.question_image.delete()
            question.question_image = None
        
        question.save()
        
        messages.success(request, 'Question updated successfully!')
        return redirect('semesters:manage_questions', subject_id=question.subject.id)
    
    return render(request, 'semesters/edit_question.html', {'question': question})

@staff_member_required
def delete_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    subject_id = question.subject.id
    
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'Question deleted successfully!')
        return redirect('semesters:manage_questions', subject_id=subject_id)
    
    return render(request, 'semesters/delete_question.html', {'question': question})

@staff_member_required
def add_question(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    
    if request.method == 'POST':
        unit = request.POST.get('unit', 1)
        question_text = request.POST.get('question_text', '')
        option_a = request.POST.get('option_a', '')
        option_b = request.POST.get('option_b', '')
        option_c = request.POST.get('option_c', '')
        option_d = request.POST.get('option_d', '')
        correct_answer = request.POST.get('correct_answer', 'A')
        explanation = request.POST.get('explanation', '')
        hint = request.POST.get('hint', '')
        difficulty = request.POST.get('difficulty', 'medium')
        tags = request.POST.get('tags', '')
        status = request.POST.get('status', 'published')
        added_by = request.POST.get('added_by', '')
        verified_by = request.POST.get('verified_by', '')
        
        # Validate all fields
        if question_text and option_a and option_b and option_c and option_d:
            question = Question.objects.create(
                subject=subject,
                unit=unit,
                question_text=question_text,
                option_a=option_a,
                option_b=option_b,
                option_c=option_c,
                option_d=option_d,
                correct_answer=correct_answer,
                explanation=explanation,
                hint=hint,
                difficulty=difficulty,
                tags=tags,
                status=status,
                added_by=added_by,
                verified_by=verified_by
            )
            
            # Handle image upload
            if 'question_image' in request.FILES:
                question.question_image = request.FILES['question_image']
                question.save()
            
            messages.success(request, 'Question added successfully!')
            return redirect('semesters:manage_questions', subject_id=subject.id)
        else:
            messages.error(request, 'All question fields and options are required!')
    
    return render(request, 'semesters/add_question.html', {'subject': subject})

@staff_member_required
def delete_unit_questions(request, subject_id, unit):
    subject = get_object_or_404(Subject, id=subject_id)
    questions = Question.objects.filter(subject=subject, unit=unit)
    question_count = questions.count()
    
    if request.method == 'POST':
        questions.delete()
        messages.success(request, f'Successfully deleted {question_count} questions from Unit {unit}!')
        return redirect('semesters:manage_questions', subject_id=subject.id)
    
    context = {
        'subject': subject,
        'unit': unit,
        'question_count': question_count,
        'questions': questions[:5]  # Show first 5 as preview
    }
    return render(request, 'semesters/delete_unit_questions.html', context)

@staff_member_required
def delete_all_questions(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    questions = Question.objects.filter(subject=subject)
    question_count = questions.count()
    
    # Get unit distribution for preview
    unit_distribution = {}
    for unit in questions.values_list('unit', flat=True).distinct().order_by('unit'):
        unit_distribution[unit] = questions.filter(unit=unit).count()
    
    if request.method == 'POST':
        questions.delete()
        messages.success(request, f'Successfully deleted all {question_count} questions from {subject.name}!')
        return redirect('semesters:manage_questions', subject_id=subject.id)
    
    context = {
        'subject': subject,
        'question_count': question_count,
        'unit_distribution': unit_distribution,
    }
    return render(request, 'semesters/delete_all_questions.html', context)


@staff_member_required
def manage_units(request, subject_id):
    """Manage units for a subject"""
    subject = get_object_or_404(Subject, id=subject_id)
    units = Unit.objects.filter(subject=subject).order_by('unit_number')
    
    # Get distinct units from questions
    question_units = Question.objects.filter(subject=subject).values_list('unit', flat=True).distinct().order_by('unit')
    
    # Get existing unit numbers
    existing_unit_numbers = set(units.values_list('unit_number', flat=True))
    
    return render(request, 'semesters/manage_units.html', {
        'subject': subject,
        'units': units,
        'question_units': question_units,
        'existing_unit_numbers': existing_unit_numbers,
    })


@staff_member_required
def create_unit(request, subject_id, unit_number):
    """Create unit information"""
    subject = get_object_or_404(Subject, id=subject_id)
    unit, created = Unit.objects.get_or_create(
        subject=subject,
        unit_number=unit_number,
        defaults={'title': '', 'topics': ''}
    )
    if created:
        messages.info(request, f'Created new Unit {unit_number}. Please add title and topics.')
    
    if request.method == 'POST':
        unit.title = request.POST.get('title', '')
        unit.topics = request.POST.get('topics', '')
        unit.save()
        messages.success(request, f'Unit {unit.unit_number} saved successfully!')
        return redirect('semesters:manage_units', subject_id=unit.subject.id)
    
    return render(request, 'semesters/edit_unit.html', {'unit': unit})


@staff_member_required
def edit_unit(request, unit_id):
    """Edit unit information"""
    unit = get_object_or_404(Unit, id=unit_id)
    
    if request.method == 'POST':
        unit.title = request.POST.get('title', '')
        unit.topics = request.POST.get('topics', '')
        unit.save()
        messages.success(request, f'Unit {unit.unit_number} saved successfully!')
        return redirect('semesters:manage_units', subject_id=unit.subject.id)
    
    return render(request, 'semesters/edit_unit.html', {'unit': unit})


# ========== PROGRAMMING QUESTION MANAGEMENT VIEWS ==========

@staff_member_required
def manage_programming_questions(request, subject_id):
    """Manage programming questions for a subject"""
    subject = get_object_or_404(Subject, id=subject_id)
    unit_filter = request.GET.get('unit', '')
    
    if unit_filter:
        programming_questions = ProgrammingQuestion.objects.filter(subject=subject, unit=unit_filter).order_by('id')
    else:
        programming_questions = ProgrammingQuestion.objects.filter(subject=subject).order_by('unit', 'id')
    
    # Get all available units for this subject
    available_units = ProgrammingQuestion.objects.filter(subject=subject).values_list('unit', flat=True).distinct().order_by('unit')
    
    context = {
        'subject': subject,
        'programming_questions': programming_questions,
        'available_units': available_units,
        'selected_unit': unit_filter,
    }
    return render(request, 'semesters/manage_programming_questions.html', context)

@staff_member_required
def add_programming_question(request, subject_id):
    """Add a single programming question"""
    subject = get_object_or_404(Subject, id=subject_id)
    
    if request.method == 'POST':
        unit = request.POST.get('unit', 1)
        question_text = request.POST.get('question_text', '')
        solution = request.POST.get('solution', '')
        hint = request.POST.get('hint', '')
        difficulty = request.POST.get('difficulty', 'medium')
        tags = request.POST.get('tags', '')
        status = request.POST.get('status', 'published')
        added_by = request.POST.get('added_by', '')
        verified_by = request.POST.get('verified_by', '')
        test_cases_raw = request.POST.get('test_cases', '')
        test_cases = None
        if test_cases_raw.strip():
            try:
                import json as _json
                test_cases = _json.loads(test_cases_raw)
                if not isinstance(test_cases, list):
                    test_cases = None
            except Exception:
                pass
        # Validate question text
        if question_text:
            programming_question = ProgrammingQuestion.objects.create(
                subject=subject,
                unit=unit,
                question_text=question_text,
                solution=solution,
                hint=hint,
                difficulty=difficulty,
                tags=tags,
                status=status,
                test_cases=test_cases,
                added_by=added_by,
                verified_by=verified_by
            )
            
            # Handle image upload
            if 'question_image' in request.FILES:
                programming_question.question_image = request.FILES['question_image']
                programming_question.save()
            # Handle CSV/data file upload (for pandas/numpy/ML questions)
            if 'csv_file' in request.FILES:
                programming_question.csv_file = request.FILES['csv_file']
                programming_question.save()
            
            messages.success(request, 'Programming question added successfully!')
            return redirect('semesters:manage_programming_questions', subject_id=subject.id)
        else:
            messages.error(request, 'Question text is required!')
    
    return render(request, 'semesters/add_programming_question.html', {'subject': subject})

@staff_member_required
def add_multiple_programming_questions(request, subject_id):
    """Add multiple programming questions manually (bulk entry) or via Excel"""
    subject = get_object_or_404(Subject, id=subject_id)
    
    if request.method == 'POST':
        submit_type = request.POST.get('submit_type', 'manual')
        
        # Handle Excel upload
        if submit_type == 'excel':
            excel_file = request.FILES.get('excel_file')
            added_by = request.POST.get('added_by', '')
            verified_by = request.POST.get('verified_by', '')
            
            if not excel_file:
                messages.error(request, 'Please select an Excel file.')
                return render(request, 'semesters/add_multiple_programming_questions.html', {'subject': subject})
            
            try:
                # Save uploaded file temporarily
                file_path = f'/tmp/{excel_file.name}' if os.name != 'nt' else f'C:\\temp\\{excel_file.name}'
                os.makedirs(os.path.dirname(file_path), exist_ok=True)
                
                with open(file_path, 'wb+') as destination:
                    for chunk in excel_file.chunks():
                        destination.write(chunk)
                
                # Read Excel file
                try:
                    df = pd.read_excel(file_path, dtype=str, keep_default_na=False)
                except Exception as e:
                    messages.error(request, f'Error reading Excel file: {str(e)}')
                    try:
                        os.remove(file_path)
                    except:
                        pass
                    return render(request, 'semesters/add_multiple_programming_questions.html', {'subject': subject})
                
                # Verify required columns
                def get_column_name(df, possible_names):
                    """Find column name with case-insensitive matching"""
                    for col in df.columns:
                        if str(col).strip().lower() in [name.lower() for name in possible_names]:
                            return col
                    return None
                
                unit_col = get_column_name(df, ['unit_number', 'unit number', 'unit'])
                question_col = get_column_name(df, ['questions', 'question', 'question_text', 'question text'])
                
                if not unit_col or not question_col:
                    messages.error(request, 'Excel file must have columns: unit_number and questions')
                    try:
                        os.remove(file_path)
                    except:
                        pass
                    return render(request, 'semesters/add_multiple_programming_questions.html', {'subject': subject})
                
                created_count = 0
                skipped_count = 0
                
                for index, row in df.iterrows():
                    try:
                        # Get unit number
                        unit_str = str(row[unit_col]).strip()
                        if not unit_str or unit_str.lower() in ['nan', 'none', '']:
                            skipped_count += 1
                            continue
                        
                        try:
                            unit = int(float(unit_str))
                        except:
                            skipped_count += 1
                            continue
                        
                        # Get question text
                        question_text = str(row[question_col]).strip()
                        if not question_text or question_text.lower() in ['nan', 'none', '']:
                            skipped_count += 1
                            continue
                        
                        # Create programming question (without solution)
                        ProgrammingQuestion.objects.create(
                            subject=subject,
                            unit=unit,
                            question_text=question_text,
                            solution='',  # No solution for Excel upload
                            added_by=added_by,
                            verified_by=verified_by
                        )
                        created_count += 1
                    except Exception as e:
                        skipped_count += 1
                        continue
                
                # Clean up temp file
                try:
                    os.remove(file_path)
                except:
                    pass
                
                if created_count > 0:
                    messages.success(request, f'Successfully added {created_count} programming question(s) from Excel!')
                    if skipped_count > 0:
                        messages.warning(request, f'Skipped {skipped_count} invalid row(s).')
                else:
                    messages.error(request, 'No valid questions were added from Excel. Please check the format.')
                
                return redirect('semesters:manage_programming_questions', subject_id=subject.id)
                
            except Exception as e:
                messages.error(request, f'Error processing Excel file: {str(e)}')
                try:
                    if 'file_path' in locals():
                        os.remove(file_path)
                except:
                    pass
                return render(request, 'semesters/add_multiple_programming_questions.html', {'subject': subject})
        
        # Handle manual entry
        unit = request.POST.get('unit', 1)
        questions_data = request.POST.get('questions_data', '')
        questions_only = request.POST.get('questions_only', '0') == '1'
        added_by = request.POST.get('added_by', '')
        verified_by = request.POST.get('verified_by', '')
        
        if questions_data:
            # Split by lines and process each question
            lines = [line.strip() for line in questions_data.split('\n') if line.strip()]
            
            created_count = 0
            skipped_count = 0
            
            current_question = None
            current_solution = None
            
            for line in lines:
                # Check if line starts with "Q:" or "Question:" to indicate a new question
                if line.lower().startswith('q:') or line.lower().startswith('question:'):
                    # Save previous question if exists
                    if current_question:
                        try:
                            ProgrammingQuestion.objects.create(
                                subject=subject,
                                unit=unit,
                                question_text=current_question.strip(),
                                solution='' if questions_only else (current_solution.strip() if current_solution else ''),
                                added_by=added_by,
                                verified_by=verified_by
                            )
                            created_count += 1
                        except Exception as e:
                            skipped_count += 1
                    
                    # Start new question
                    current_question = line.split(':', 1)[1].strip() if ':' in line else line[2:].strip()
                    current_solution = None
                # Check if line starts with "S:" or "Solution:" to indicate solution (skip if questions_only)
                elif (line.lower().startswith('s:') or line.lower().startswith('solution:')) and not questions_only:
                    current_solution = line.split(':', 1)[1].strip() if ':' in line else line[2:].strip()
                else:
                    # Append to current question or solution
                    if current_solution is not None:
                        # Append to solution
                        current_solution += '\n' + line
                    else:
                        # Append to question
                        if current_question:
                            current_question += '\n' + line
                        else:
                            current_question = line
            
            # Save last question
            if current_question:
                try:
                    ProgrammingQuestion.objects.create(
                        subject=subject,
                        unit=unit,
                        question_text=current_question.strip(),
                        solution=current_solution.strip() if current_solution else '',
                        added_by=added_by,
                        verified_by=verified_by
                    )
                    created_count += 1
                except Exception as e:
                    skipped_count += 1
            
            if created_count > 0:
                messages.success(request, f'Successfully added {created_count} programming question(s)!')
                if skipped_count > 0:
                    messages.warning(request, f'Skipped {skipped_count} invalid question(s).')
            else:
                messages.error(request, 'No valid questions were added. Please check the format.')
            
            return redirect('semesters:manage_programming_questions', subject_id=subject.id)
        else:
            messages.error(request, 'Please enter at least one question!')
    
    return render(request, 'semesters/add_multiple_programming_questions.html', {'subject': subject})

@staff_member_required
def edit_programming_question(request, programming_question_id):
    """Edit a programming question"""
    programming_question = get_object_or_404(ProgrammingQuestion, id=programming_question_id)
    
    if request.method == 'POST':
        programming_question.unit = request.POST.get('unit', programming_question.unit)
        programming_question.question_text = request.POST.get('question_text', programming_question.question_text)
        programming_question.solution = request.POST.get('solution', programming_question.solution)
        programming_question.hint = request.POST.get('hint', programming_question.hint)
        programming_question.difficulty = request.POST.get('difficulty', programming_question.difficulty)
        programming_question.tags = request.POST.get('tags', programming_question.tags)
        programming_question.status = request.POST.get('status', programming_question.status)
        programming_question.added_by = request.POST.get('added_by', programming_question.added_by)
        programming_question.verified_by = request.POST.get('verified_by', programming_question.verified_by)
        tc_raw = request.POST.get('test_cases', '')
        if tc_raw.strip():
            try:
                import json as _json
                programming_question.test_cases = _json.loads(tc_raw) if isinstance(_json.loads(tc_raw), list) else programming_question.test_cases
            except Exception:
                pass
        elif request.POST.get('clear_test_cases') == 'on':
            programming_question.test_cases = None
        # Handle image upload
        if 'question_image' in request.FILES:
            programming_question.question_image = request.FILES['question_image']
        
        # Handle image removal
        if request.POST.get('remove_image') == 'on' and programming_question.question_image:
            programming_question.question_image.delete()
            programming_question.question_image = None
        
        # Handle CSV file upload
        if 'csv_file' in request.FILES:
            if programming_question.csv_file:
                programming_question.csv_file.delete(save=False)
            programming_question.csv_file = request.FILES['csv_file']
        # Handle CSV removal
        if request.POST.get('remove_csv') == 'on' and programming_question.csv_file:
            programming_question.csv_file.delete(save=False)
            programming_question.csv_file = None
        
        programming_question.save()
        
        messages.success(request, 'Programming question updated successfully!')
        # Preserve unit filter if it was set
        unit = request.GET.get('unit', '') or programming_question.unit
        url = reverse('semesters:manage_programming_questions', args=[programming_question.subject.id])
        if unit:
            url += f'?unit={unit}'
        return redirect(url)
    
    import json as _json
    test_cases_json = _json.dumps(programming_question.test_cases, indent=2) if programming_question.test_cases else ''
    return render(request, 'semesters/edit_programming_question.html', {
        'programming_question': programming_question,
        'test_cases_json': test_cases_json,
    })

@staff_member_required
def delete_programming_question(request, programming_question_id):
    """Delete a programming question"""
    programming_question = get_object_or_404(ProgrammingQuestion, id=programming_question_id)
    subject_id = programming_question.subject.id
    
    if request.method == 'POST':
        # Get unit from request or from the question being deleted
        unit = request.GET.get('unit', '') or programming_question.unit
        programming_question.delete()
        messages.success(request, 'Programming question deleted successfully!')
        # Preserve unit filter
        url = reverse('semesters:manage_programming_questions', args=[subject_id])
        if unit:
            url += f'?unit={unit}'
        return redirect(url)
    
    return render(request, 'semesters/delete_programming_question.html', {'programming_question': programming_question})

@staff_member_required
def delete_unit_programming_questions(request, subject_id, unit):
    """Delete all programming questions in a unit"""
    subject = get_object_or_404(Subject, id=subject_id)
    programming_questions = ProgrammingQuestion.objects.filter(subject=subject, unit=unit)
    question_count = programming_questions.count()
    
    if request.method == 'POST':
        programming_questions.delete()
        messages.success(request, f'Successfully deleted {question_count} programming questions from Unit {unit}!')
        # Preserve unit filter
        url = reverse('semesters:manage_programming_questions', args=[subject.id])
        url += f'?unit={unit}'
        return redirect(url)
    
    context = {
        'subject': subject,
        'unit': unit,
        'question_count': question_count,
        'programming_questions': programming_questions[:5]  # Show first 5 as preview
    }
    return render(request, 'semesters/delete_unit_programming_questions.html', context)

@staff_member_required
def delete_all_programming_questions(request, subject_id):
    """Delete all programming questions for a subject"""
    subject = get_object_or_404(Subject, id=subject_id)
    programming_questions = ProgrammingQuestion.objects.filter(subject=subject)
    question_count = programming_questions.count()
    
    # Get unit distribution for preview
    unit_distribution = {}
    for unit in programming_questions.values_list('unit', flat=True).distinct().order_by('unit'):
        unit_distribution[unit] = programming_questions.filter(unit=unit).count()
    
    if request.method == 'POST':
        programming_questions.delete()
        messages.success(request, f'Successfully deleted all {question_count} programming questions from {subject.name}!')
        return redirect('semesters:manage_programming_questions', subject_id=subject.id)
    
    context = {
        'subject': subject,
        'question_count': question_count,
        'unit_distribution': unit_distribution,
    }
    return render(request, 'semesters/delete_all_programming_questions.html', context)


@staff_member_required
def programming_question_analytics(request):
    """View analytics for programming question module access"""
    # Get filter parameters
    subject_id = request.GET.get('subject', '')
    unit = request.GET.get('unit', '')
    days = int(request.GET.get('days', 30))  # Default: last 30 days
    
    # Base queryset
    accesses = ProgrammingQuestionAccess.objects.all()
    
    # Apply filters
    if subject_id:
        accesses = accesses.filter(subject_id=subject_id)
    if unit:
        accesses = accesses.filter(unit=unit)
    
    # Date filter
    date_from = timezone.now() - timedelta(days=days)
    accesses = accesses.filter(accessed_at__gte=date_from)
    
    # Get all subjects for filter dropdown
    subjects = Subject.objects.all().order_by('semester', 'name')
    
    # Statistics
    total_accesses = accesses.count()
    unique_users = accesses.values('user').distinct().count()
    avg_per_user = round(total_accesses / unique_users, 1) if unique_users > 0 else 0
    
    # Access by subject
    access_by_subject = accesses.values('subject__name', 'subject__semester__name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Access by unit (for selected subject or all)
    if subject_id:
        access_by_unit = accesses.filter(subject_id=subject_id).values('unit').annotate(
            count=Count('id')
        ).order_by('unit')
    else:
        access_by_unit = accesses.values('subject__name', 'unit').annotate(
            count=Count('id')
        ).order_by('subject__name', 'unit')
    
    # Access by user
    access_by_user = accesses.values('user__username', 'user__email').annotate(
        count=Count('id')
    ).order_by('-count')[:20]  # Top 20 users
    
    # Recent accesses
    recent_accesses = accesses.select_related('user', 'subject').order_by('-accessed_at')[:50]
    
    # Daily access trend (last 30 days)
    daily_trend = []
    for i in range(days):
        date = timezone.now() - timedelta(days=i)
        count = accesses.filter(
            accessed_at__date=date.date()
        ).count()
        daily_trend.append({
            'date': date.date(),
            'count': count
        })
    daily_trend.reverse()  # Show oldest to newest
    
    context = {
        'subjects': subjects,
        'selected_subject_id': subject_id,
        'selected_unit': unit,
        'selected_days': days,
        'total_accesses': total_accesses,
        'unique_users': unique_users,
        'avg_per_user': avg_per_user,
        'access_by_subject': access_by_subject,
        'access_by_unit': access_by_unit,
        'access_by_user': access_by_user,
        'recent_accesses': recent_accesses,
        'daily_trend': daily_trend,
    }
    return render(request, 'semesters/programming_question_analytics.html', context)
