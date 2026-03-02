#!/usr/bin/env python
"""
Fix Corrupted Excel File Script
This script attempts to repair and extract data from corrupted Excel files.
"""

import sys
import os
import pandas as pd
import openpyxl
from pathlib import Path

def fix_excel_file(input_file, output_file=None):
    """
    Attempts to fix a corrupted Excel file by trying multiple reading strategies.
    
    Args:
        input_file: Path to corrupted Excel file
        output_file: Path to save fixed file (optional)
    """
    if not os.path.exists(input_file):
        print(f"❌ Error: File '{input_file}' not found!")
        return False
    
    print(f"🔧 Attempting to fix: {input_file}")
    print("=" * 60)
    
    # Try pandas with multiple engines
    print("\n1️⃣ Trying pandas with openpyxl engine...")
    try:
        df = pd.read_excel(input_file, dtype=str, keep_default_na=False, engine='openpyxl')
        print(f"   ✅ Success! Read {len(df)} rows")
        success = True
    except Exception as e:
        print(f"   ❌ Failed: {str(e)[:100]}")
        success = False
    
    if not success:
        print("\n2️⃣ Trying pandas with xlrd engine...")
        try:
            df = pd.read_excel(input_file, dtype=str, keep_default_na=False, engine='xlrd')
            print(f"   ✅ Success! Read {len(df)} rows")
            success = True
        except Exception as e:
            print(f"   ❌ Failed: {str(e)[:100]}")
            success = False
    
    if not success:
        print("\n3️⃣ Trying openpyxl read_only mode...")
        try:
            wb = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
            ws = wb.active
            
            data_rows = []
            headers = None
            
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx == 0:
                    headers = [str(cell) if cell is not None else '' for cell in row]
                else:
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    if any(row_data):
                        data_rows.append(row_data)
            
            if headers and data_rows:
                df = pd.DataFrame(data_rows, columns=headers)
                print(f"   ✅ Success! Read {len(df)} rows using cell-by-cell method")
                success = True
            else:
                print("   ❌ No data found")
                success = False
            
            wb.close()
        except Exception as e:
            print(f"   ❌ Failed: {str(e)[:100]}")
            success = False
    
    if not success:
        print("\n4️⃣ Trying openpyxl normal mode...")
        try:
            wb = openpyxl.load_workbook(input_file, read_only=False, data_only=True)
            ws = wb.active
            
            data_rows = []
            headers = None
            
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx == 0:
                    headers = [str(cell) if cell is not None else '' for cell in row]
                else:
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    if any(row_data):
                        data_rows.append(row_data)
            
            if headers and data_rows:
                df = pd.DataFrame(data_rows, columns=headers)
                print(f"   ✅ Success! Read {len(df)} rows")
                success = True
            else:
                print("   ❌ No data found")
                success = False
            
            wb.close()
        except Exception as e:
            print(f"   ❌ Failed: {str(e)[:100]}")
            success = False
    
    if success:
        # Save to new file
        if output_file is None:
            input_path = Path(input_file)
            output_file = input_path.parent / f"{input_path.stem}_FIXED.xlsx"
        
        print(f"\n💾 Saving fixed file to: {output_file}")
        try:
            df.to_excel(output_file, index=False, engine='openpyxl')
            print(f"   ✅ Successfully saved fixed file!")
            print(f"\n✅ File fixed! You can now upload '{output_file}' to the system.")
            return True
        except Exception as e:
            print(f"   ❌ Error saving file: {str(e)}")
            print(f"\n⚠️  Data was read but couldn't be saved. Try manual copy-paste method.")
            return False
    else:
        print("\n" + "=" * 60)
        print("❌ All automatic repair methods failed!")
        print("\n🔧 MANUAL FIX REQUIRED:")
        print("1. Open the file in Microsoft Excel")
        print("2. Select all (Ctrl+A) and Copy (Ctrl+C)")
        print("3. Create a NEW blank Excel workbook")
        print("4. Paste (Ctrl+V)")
        print("5. Save as new .xlsx file")
        print("6. Upload the new file")
        return False

def main():
    if len(sys.argv) < 2:
        print("Usage: python fix_corrupted_excel.py <input_file.xlsx> [output_file.xlsx]")
        print("\nExample:")
        print("  python fix_corrupted_excel.py corrupted_file.xlsx")
        print("  python fix_corrupted_excel.py corrupted_file.xlsx fixed_file.xlsx")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    success = fix_excel_file(input_file, output_file)
    sys.exit(0 if success else 1)

if __name__ == '__main__':
    main()

